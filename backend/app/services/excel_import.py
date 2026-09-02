"""Importeert de bestaande Excel-planning (schema seizoen ...xlsx) in de database.

Verwacht een werkblad met een headerrij die o.a. bevat: wedstrijdnummer, datum,
thuisteam, uitteam, locatie, gevolgd door één kolom per speler (met waarden
v/x/?/1, zie EXCEL_AVAILABILITY_MAP), en optioneel "aantal beschikbare spelers"
als laatste kolom. Zie functioneel ontwerp v1, secties 4 en 5.
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.availability import Availability
from app.models.enums import EXCEL_AVAILABILITY_MAP, AvailabilityStatus, MatchType
from app.models.match import Match
from app.models.player import Player

KNOWN_COLUMNS = {
    "wedstrijdnummer": "nummer",
    "nummer": "nummer",
    "datum": "datum",
    "thuisteam": "thuisteam",
    "uitteam": "uitteam",
    "locatie": "locatie",
}
IGNORED_COLUMNS = {"aantal beschikbare spelers", "aantal", "opmerking", "opmerkingen"}


def _match_type_from_nummer(nummer: str | None) -> MatchType:
    if not nummer:
        return MatchType.COMPETITIE
    prefix = nummer.strip()[:1].upper()
    if prefix == "B":
        return MatchType.BEKER
    if prefix == "I":
        return MatchType.INHAAL
    return MatchType.COMPETITIE


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def import_excel(db: Session, file_bytes: bytes) -> dict:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"matches_created": 0, "matches_updated": 0, "players_created": 0, "availability_upserted": 0}

    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]

    column_roles: list[str] = []
    player_columns: list[tuple[int, str]] = []
    for idx, col_name in enumerate(header):
        if col_name in KNOWN_COLUMNS:
            column_roles.append(KNOWN_COLUMNS[col_name])
        elif col_name in IGNORED_COLUMNS or col_name == "":
            column_roles.append("ignore")
        else:
            column_roles.append("player")
            player_columns.append((idx, header[idx]))

    player_cache: dict[str, Player] = {p.naam.lower(): p for p in db.query(Player).all()}
    players_created = 0

    def get_or_create_player(naam: str) -> Player:
        nonlocal players_created
        key = naam.lower()
        if key in player_cache:
            return player_cache[key]
        player = Player(naam=naam)
        db.add(player)
        db.flush()
        player_cache[key] = player
        players_created += 1
        return player

    matches_created = 0
    matches_updated = 0
    availability_upserted = 0

    for raw_row in rows[1:]:
        if raw_row is None or all(cell is None for cell in raw_row):
            continue

        data: dict[str, object] = {}
        player_values: list[tuple[str, object]] = []
        for idx, role in enumerate(column_roles):
            value = raw_row[idx] if idx < len(raw_row) else None
            if role == "ignore":
                continue
            if role == "player":
                player_naam = header[idx].strip().title()
                player_values.append((player_naam, value))
            else:
                data[role] = value

        datum_raw = data.get("datum")
        thuisteam = str(data.get("thuisteam") or "").strip()
        uitteam = str(data.get("uitteam") or "").strip()
        if not datum_raw or not thuisteam or not uitteam:
            continue

        if isinstance(datum_raw, datetime):
            datum: date = datum_raw.date()
        elif isinstance(datum_raw, date):
            datum = datum_raw
        else:
            continue

        nummer = str(data.get("nummer") or "").strip() or None
        locatie = str(data.get("locatie") or "").strip() or None
        external_id = nummer or f"{datum.isoformat()}-{thuisteam}-{uitteam}"

        match = db.query(Match).filter(Match.external_id == external_id).first()
        if match:
            match.datum = datum
            match.thuisteam = thuisteam
            match.uitteam = uitteam
            match.locatie = locatie
            match.nummer = nummer
            match.type = _match_type_from_nummer(nummer)
            matches_updated += 1
        else:
            match = Match(
                external_id=external_id,
                datum=datum,
                thuisteam=thuisteam,
                uitteam=uitteam,
                locatie=locatie,
                nummer=nummer,
                type=_match_type_from_nummer(nummer),
            )
            db.add(match)
            db.flush()
            matches_created += 1

        for player_naam, raw_value in player_values:
            player = get_or_create_player(player_naam)
            status: AvailabilityStatus = EXCEL_AVAILABILITY_MAP.get(
                _normalize(raw_value), AvailabilityStatus.NO_RESPONSE
            )

            availability = (
                db.query(Availability)
                .filter(Availability.match_id == match.id, Availability.player_id == player.id)
                .first()
            )
            if availability:
                availability.status = status
            else:
                db.add(Availability(match_id=match.id, player_id=player.id, status=status))
            availability_upserted += 1

    db.commit()

    return {
        "matches_created": matches_created,
        "matches_updated": matches_updated,
        "players_created": players_created,
        "availability_upserted": availability_upserted,
    }
